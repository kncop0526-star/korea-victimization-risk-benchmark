# -*- coding: utf-8 -*-
"""build_human_reread_sheet.py (reviewer M3 / 4-reviewer convergent demand) — break the all-synthetic
round-trip with HUMAN raters. Builds a STRATIFIED 200-persona sample (balanced over age_band x sex),
strips leaked '(속성 N)' parentheticals, and emits THREE blank rater sheets so 3 independent coders
decode FV/DL/AD/SI (1-5) from narrative alone. --score reads the 3 filled sheets and reports
inter-rater reliability (mean pairwise quadratic-weighted kappa AND ordinal Krippendorff alpha) plus
the headline number: 3-rater consensus vs the sampler's value (human-vs-sampled QWK).

USAGE (Daniel + 2 colleagues):
  py src/build_human_reread_sheet.py --build --in data/processed/enriched_stage23_N2000.jsonl \
        --parts data/processed/enriched_1M_v4_parts --n 200 --out results
    -> results/human_reread_rater1.xlsx, _rater2.xlsx, _rater3.xlsx  (one per coder; identical personas)
    -> results/human_reread_key.csv  (KEEP PRIVATE: sampled values)
  # each coder fills their own sheet's human_* columns (1-5), independently, blind to the others
  py src/build_human_reread_sheet.py --score --sheets results/human_reread_rater1_FILLED.xlsx \
        results/human_reread_rater2_FILLED.xlsx results/human_reread_rater3_FILLED.xlsx \
        --key results/human_reread_key.csv --out results
"""
from __future__ import annotations
import argparse, json, re, itertools
from pathlib import Path
import numpy as np, pandas as pd

ATTRS=["financial_vulnerability","digital_literacy","authority_deference","social_isolation"]
KOR={"financial_vulnerability":"재정 취약성","digital_literacy":"디지털 활용",
     "authority_deference":"권위 순응","social_isolation":"사회적 고립"}
LEAK=re.compile(r'\s*\((재정 취약성|디지털 활용|권위 순응|사회적 고립)\s*[1-5]\)')

def qwk(a,b):
    a=np.asarray(a,float); b=np.asarray(b,float); m=~(np.isnan(a)|np.isnan(b)); a,b=a[m],b[m]
    if len(a)<2: return float("nan")
    cats=sorted(set(a)|set(b)); idx={c:i for i,c in enumerate(cats)}; k=len(cats)
    if k<2: return float("nan")
    O=np.zeros((k,k))
    for x,y in zip(a,b): O[idx[x],idx[y]]+=1
    W=np.array([[((i-j)**2)/((k-1)**2) for j in range(k)] for i in range(k)])
    r=O.sum(1); c=O.sum(0); E=np.outer(r,c)/O.sum()
    num=(W*O).sum(); den=(W*E).sum()
    return float(1-num/den) if den>0 else float("nan")

def kripp_alpha_ordinal(M):
    """M: (units x raters) array of ratings 1-5 with np.nan for missing. Ordinal Krippendorff alpha."""
    vals=[v for v in np.unique(M[~np.isnan(M)])]
    if len(vals)<2: return float("nan")
    vmap={v:i for i,v in enumerate(sorted(vals))}; K=len(vmap)
    # coincidence matrix
    o=np.zeros((K,K)); 
    for row in M:
        r=row[~np.isnan(row)]
        mu=len(r)
        if mu<2: continue
        for i in range(mu):
            for j in range(mu):
                if i!=j: o[vmap[r[i]],vmap[r[j]]] += 1.0/(mu-1)
    n_marg=o.sum(0); n=n_marg.sum()
    if n<1: return float("nan")
    # ordinal metric delta^2(c,k) = (sum_{g=c..k} n_g - (n_c+n_k)/2)^2
    order=sorted(vmap, key=lambda v:vmap[v]); 
    def d2(ci,ki):
        lo,hi=min(ci,ki),max(ci,ki)
        s=sum(n_marg[g] for g in range(lo,hi+1)) - (n_marg[ci]+n_marg[ki])/2.0
        return s*s
    Do=sum(o[c,k]*d2(c,k) for c in range(K) for k in range(K))
    De=sum(n_marg[c]*n_marg[k]*d2(c,k) for c in range(K) for k in range(K))/(n-1) if n>1 else 0
    return float(1-Do/De) if De>0 else float("nan")

def build(inp, parts, n, out, coders=3):
    rows=[json.loads(l) for l in open(inp,encoding="utf-8") if l.strip()]
    # join demographics via uuid from parquet for stratification
    import glob
    dem=pd.concat([pd.read_parquet(p,columns=["uuid","age_band","sex_mf"]) for p in sorted(glob.glob(parts+"/*.parquet"))],ignore_index=True)
    dmap=dem.set_index("uuid")[["age_band","sex_mf"]].to_dict("index")
    df=pd.DataFrame([{"i":i,"uuid":r.get("uuid"),
                      "age_band":dmap.get(r.get("uuid"),{}).get("age_band","?"),
                      "sex_mf":dmap.get(r.get("uuid"),{}).get("sex_mf","?")} for i,r in enumerate(rows)])
    # stratified: equal per (age_band x sex) cell
    rng=np.random.default_rng(42); cells=df.groupby(["age_band","sex_mf"])
    per=max(1, n//max(1,cells.ngroups)); sel=[]
    for _,g in cells: sel += list(rng.choice(g["i"].values, min(per,len(g)), replace=False))
    sel=sel[:n]
    sheet=[]; key=[]
    for k,i in enumerate(sel):
        d=rows[i]; narr=LEAK.sub("", d.get("attr_narrative","") or "").strip(); pid=d.get("uuid",f"P{k:04d}")
        sheet.append({"persona_id":pid,"narrative":narr}); a=d.get("attr",{})
        key.append({"persona_id":pid, **{f"sampled_{x}":a.get(x) for x in ATTRS}})
    out=Path(out); out.mkdir(parents=True,exist_ok=True)
    import openpyxl
    cols=["persona_id","narrative"]+[f"human_{x}" for x in ATTRS]
    INSTR=[
      ["[\ud3c9\uc815 \uc9c0\uce68] \uc778\ubb3c \uc790\uae30\uc18c\uac1c(narrative)\ub9cc \uc77d\uace0 \ub124 \uac00\uc9c0 \ud2b9\uc131\uc744 \uac01\uac01 1~5\uc810\uc73c\ub85c \ucc44\uc810\ud558\uc138\uc694."],
      ["\ucc44\uc810\uce78: C=\uc7ac\uc815\ucde8\uc57d\uc131  D=\ub514\uc9c0\ud138\ud65c\uc6a9  E=\uad8c\uc704\uc21c\uc751  F=\uc0ac\ud68c\uc801\uace0\ub9bd  (\ube48\uce78 \uc5c6\uc774 \ubaa8\ub450 1~5 \uc785\ub825)"],
      [""],
      ["\u25a0 \uc7ac\uc815\ucde8\uc57d\uc131(financial_vulnerability): 1=\uacbd\uc81c\uc801\uc73c\ub85c \ub9e4\uc6b0 \uc548\uc815 \u2026 5=\uc218\uc785 \uc5c6\uc74c/\uc0dd\uacc4 \uacf3\ub780"],
      ["\u25a0 \ub514\uc9c0\ud138\ud65c\uc6a9(digital_literacy): 1=\ub514\uc9c0\ud138 \uae30\uae30 \uac70\uc758 \ubabb \ub2e4\ub8e8\uba70 \u2026 5=\ub2a5\uc219\ud558\uac8c \uc790\uc720\uc790\uc7ac"],
      ["\u25a0 \uad8c\uc704\uc21c\uc751(authority_deference): 1=\ube44\ud310\uc801/\ub0a9\ub4dd\ud574\uc57c \ub530\ub984 \u2026 5=\uc717\uc0ac\ub78c\u00b7\uad00\uc2b5\uc5d0 \ubb34\uc870\uac74 \uc21c\uc751"],
      ["\u25a0 \uc0ac\ud68c\uc801\uace0\ub9bd(social_isolation): 1=\uad50\ub958 \ud65c\ubc1c/\ub3c4\uc6c0\ubc1b\uc744 \uc0ac\ub78c \ub9ce\uc74c \u2026 5=\uc644\uc804\ud788 \uace0\ub9bd/\uc758\uc9c0\ud560 \uc0ac\ub78c \uc5c6\uc74c"],
      [""],
      ["\uaddc\uce59: \u2460 \ub2e4\ub978 \ud3c9\uc815\uc790\uc640 \uc0c1\uc758 \uae08\uc9c0(\ub3c5\ub9bd \ucc44\uc810) \u2461 \uc815\ub2f5\ud45c\ub97c \ubcf4\uc9c0 \ub9d0 \uac83 \u2462 \uae00\uc5d0 \ub2e8\uc11c\uac00 \uc57d\ud558\uba74 \ubcf8\uc778 \ud310\ub2e8\uc73c\ub85c \uacb0\uc815"],
      ["\uc608\uc2dc: '\uc218\uc785\uc774 \uc5c6\uc5b4 \uc0dd\ud65c\ube44 \uac71\uc815' \u2192 \uc7ac\uc815\ucde8\uc57d\uc131 \ub192\uc74c / '\ubcf5\uc7a1\ud55c \uae30\uae30\ub294 \ubabb \ub2e4\ub8e8' \u2192 \ub514\uc9c0\ud138\ud65c\uc6a9 \ub0ae\uc74c /"],
      ["      '\ub0a9\ub4dd\ub418\uc5b4\uc57c \uc6c0\uc9c1\uc784' \u2192 \uad8c\uc704\uc21c\uc751 \ub0ae\uc74c / '\uce5c\uad6c\ub4e4\uacfc \uc790\uc8fc \uad50\ub958' \u2192 \uc0ac\ud68c\uc801\uace0\ub9bd \ub0ae\uc74c"],
      [""],
      ["\uc791\uc131 \ud6c4 \ud30c\uc77c\uba85\uc744 human_reread_raterN_FILLED.xlsx \ub85c \uc800\uc7a5\ud574 \ub2f4\ub2f9\uc790\uc5d0\uac8c \uc804\ub2ec\ud558\uc138\uc694."],
    ]
    labels=["", "\u2190 \uc774 \uae00\ub9cc \uc77d\uace0 \ucc44\uc810", "\uc7ac\uc815\ucde8\uc57d\uc131 1-5","\ub514\uc9c0\ud138\ud65c\uc6a9 1-5","\uad8c\uc704\uc21c\uc751 1-5","\uc0ac\ud68c\uc801\uace0\ub9bd 1-5"]
    for rno in range(1,coders+1):
        order=list(range(len(sheet)))
        np.random.default_rng(100+rno).shuffle(order)   # independent persona order per coder (order/cross-ref effects)
        wb=openpyxl.Workbook()
        gi=wb.active; gi.title="\uc9c0\uce68"
        for row in INSTR: gi.append(row)
        gi.column_dimensions["A"].width=110
        ws=wb.create_sheet("rater"); ws.append(cols); ws.append(labels)
        for oi in order:
            r=sheet[oi]; ws.append([r["persona_id"],r["narrative"],"","","",""])
        ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=95
        for col in ("C","D","E","F"): ws.column_dimensions[col].width=14
        ws.freeze_panes="C3"
        wb.save(out/f"human_reread_subset{n}_coder{rno}.xlsx")
    pd.DataFrame(key).to_csv(out/"human_reread_key.csv",index=False,encoding="utf-8-sig")
    print(f"[ok] {coders} coder sheets (human_reread_subset{n}_coder1..{coders}.xlsx; {len(sheet)} stratified personas, independent shuffle, leak-stripped) + human_reread_key.csv (PRIVATE)")
    print(f"[strata] per (age_band x sex) cell ~{per}; cells={cells.ngroups}")

def score(sheets, key, out):
    import openpyxl
    K=pd.read_csv(key); res=[]
    R={}
    for si,sh in enumerate(sheets,1):
        wb=openpyxl.load_workbook(sh); ws=wb["rater"]; rows=list(ws.values); hdr=list(rows[0])
        data=[dict(zip(hdr,r)) for r in rows[1:] if r[0] and r[0]!=""]
        R[si]=pd.DataFrame(data)
    for x in ATTRS:
        # consensus = mean across raters (rounded), per persona
        merged=K[["persona_id",f"sampled_{x}"]].copy()
        cols=[]
        for si,H in R.items():
            h=pd.to_numeric(H.set_index("persona_id")[f"human_{x}"],errors="coerce")
            merged[f"r{si}"]=merged["persona_id"].map(h); cols.append(f"r{si}")
        Mraw=merged[cols].to_numpy(float)
        # inter-rater: mean pairwise QWK + Krippendorff ordinal alpha
        pw=[qwk(merged[a],merged[b]) for a,b in itertools.combinations(cols,2)]
        meanpw=float(np.nanmean(pw)) if pw else float("nan")
        alpha=kripp_alpha_ordinal(Mraw)
        # consensus vs sampled
        cons=np.round(np.nanmean(Mraw,axis=1))
        v=qwk(cons, pd.to_numeric(merged[f"sampled_{x}"],errors="coerce"))
        res.append({"attr":x,"interrater_meanQWK":round(meanpw,3),"krippendorff_alpha":round(alpha,3),
                    "consensus_vs_sampled_QWK":round(v,3),"n":int(np.isfinite(cons).sum())})
    cmp=pd.DataFrame(res); Path(out).mkdir(parents=True,exist_ok=True)
    cmp.to_csv(Path(out)/"human_reread_results.csv",index=False,encoding="utf-8-sig")
    print(cmp.to_string(index=False)); print("\n[ok] results/human_reread_results.csv")

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--build",action="store_true"); ap.add_argument("--score",action="store_true")
    ap.add_argument("--in",dest="inp",default="data/processed/enriched_stage23_N2000.jsonl")
    ap.add_argument("--parts",default="data/processed/enriched_1M_v4_parts")
    ap.add_argument("--n",type=int,default=200); ap.add_argument("--coders",type=int,default=3); ap.add_argument("--out",default="results")
    ap.add_argument("--sheets",nargs="+"); ap.add_argument("--key")
    a=ap.parse_args()
    if a.build: build(a.inp,a.parts,a.n,a.out,a.coders)
    elif a.score:
        if not (a.sheets and a.key): raise SystemExit("[fatal] --score needs --sheets s1 s2 s3 --key")
        score(a.sheets,a.key,a.out)
    else: raise SystemExit("[fatal] use --build or --score")
if __name__=="__main__": main()
