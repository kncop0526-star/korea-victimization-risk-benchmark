"""
build_face_validity_sheet.py — Technical Validation 4.4 prep (face validity review sheet).

After the pinned-LLM Stage-2/3 pass produces real narratives, this samples N records and writes a
blind-scoring sheet for an expert: the narrative plus the sampled attribute levels, with empty
columns for a plausibility score and a consistency judgment. The reviewer reads each narrative and
rates whether it is a believable Korean persona description and whether it matches the fixed levels.
Aggregating the sheet fills §4.4.

By default it refuses to run on stub narratives (no value in reviewing the deterministic stub); pass
--allow-stub only for a layout preview.

Outputs
  results/face_validity_sheet.csv   (utf-8-sig, Excel-ready) — one row per sampled persona
  results/face_validity_sheet.md    (same content, readable in a viewer)

Usage (after the real pass)
  python src/build_face_validity_sheet.py --in data/processed/enriched_stage23.jsonl --n 50 --seed 7
"""
from __future__ import annotations
import argparse, json, random
from pathlib import Path

ATTRS = ["financial_vulnerability", "digital_literacy", "authority_deference", "social_isolation"]


def load(path: str):
    recs, stub = [], False
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            if "STUB" in str((r.get("gen_meta") or {}).get("actual_model", "")):
                stub = True
            recs.append(r)
    return recs, stub


def build_xlsx(sample, out_path):
    """Write a readable scoring workbook: narrative wraps in-cell, score columns with dropdowns."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("[skip] openpyxl not installed (pip install openpyxl) — xlsx not written.")
        return False
    wb = Workbook(); ws = wb.active; ws.title = "face_validity"
    headers = ["idx", "uuid", "levels (FV/DL/AD/SI)", "narrative",
               "plausible_1to5", "consistent_Y_N", "notes"]
    ws.append(headers)
    hd = Font(bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="4C72B0")
    thin = Side(style="thin", color="CCCCCC"); border = Border(*([thin]*4))
    for c in range(1, len(headers)+1):
        cell = ws.cell(1, c); cell.font = hd; cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = border
    widths = [5, 12, 16, 78, 14, 14, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w
    top = Alignment(vertical="top", wrap_text=True)
    for r, x in enumerate(sample, start=2):
        attr = x.get("attr", {})
        lv = f"FV{attr.get('financial_vulnerability')} DL{attr.get('digital_literacy')} " \
             f"AD{attr.get('authority_deference')} SI{attr.get('social_isolation')}"
        narr = x.get("attr_narrative", "")
        ws.append([r-1, str(x.get("uuid",""))[:8], lv, narr, "", "", ""])
        for c in range(1, len(headers)+1):
            cell = ws.cell(r, c); cell.alignment = top; cell.border = border
        # estimate wrapped height (~40 Korean chars per line at width 78)
        lines = max(3, (len(narr) // 38) + narr.count("\n") + 1)
        ws.row_dimensions[r].height = 15 * lines
    n = len(sample) + 1
    dv_p = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_c = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_p); ws.add_data_validation(dv_c)
    dv_p.add(f"E2:E{n}"); dv_c.add(f"F2:F{n}")
    ws.freeze_panes = "A2"
    wb.save(out_path)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", default="data/processed/enriched_stage23.jsonl")
    ap.add_argument("--out", default="results")
    ap.add_argument("--n", type=int, default=50)
    ap.add_argument("--seed", type=int, default=7)
    ap.add_argument("--allow-stub", action="store_true")
    args = ap.parse_args()

    recs, stub = load(args.inp)
    if stub and not args.allow_stub:
        raise SystemExit("Input is STUB narratives. Run the pinned-LLM pass first, "
                         "or pass --allow-stub for a layout preview only.")
    random.seed(args.seed)
    sample = random.sample(recs, min(args.n, len(recs)))

    out = Path(args.out); out.mkdir(parents=True, exist_ok=True)
    import csv
    cols = ["idx", "uuid"] + [f"level_{a}" for a in ATTRS] + [
        "narrative", "plausible_1to5", "consistent_with_levels_Y_N", "reviewer_notes"]
    with open(out / "face_validity_sheet.csv", "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f); w.writerow(cols)
        for i, r in enumerate(sample, 1):
            attr = r.get("attr", {})
            w.writerow([i, r.get("uuid")] + [attr.get(a) for a in ATTRS] +
                       [r.get("attr_narrative", ""), "", "", ""])

    with open(out / "face_validity_sheet.md", "w", encoding="utf-8") as f:
        f.write(f"# Face-validity review sheet (N={len(sample)})\n\n")
        f.write("Score each narrative: **plausible_1to5** (1=implausible, 5=fully believable), "
                "**consistent** (does the text match the fixed levels? Y/N), notes optional.\n\n")
        f.write("Model: " + str((sample[0].get('gen_meta') or {}).get('actual_model', '?')) +
                ("  **[STUB PREVIEW — not for real scoring]**\n\n" if stub else "\n\n"))
        for i, r in enumerate(sample, 1):
            attr = r.get("attr", {})
            lv = ", ".join(f"{a.split('_')[0]}={attr.get(a)}" for a in ATTRS)
            f.write(f"## {i}. `{r.get('uuid','')[:8]}`  ({lv})\n\n")
            f.write(f"> {r.get('attr_narrative','')}\n\n")
            f.write("plausible (1-5): ___   consistent (Y/N): ___   notes: \n\n---\n\n")

    xlsx_ok = build_xlsx(sample, out / "face_validity_sheet.xlsx")

    tag = "  [STUB PREVIEW]" if stub else ""
    xtra = " (+ .md, .xlsx)" if xlsx_ok else " (+ .md)"
    print(f"[ok] face-validity sheet for {len(sample)} personas -> "
          f"{out/'face_validity_sheet.csv'}{xtra}{tag}")
    print("[next] expert fills plausible_1to5 + consistent_Y_N; aggregate to report §4.4 "
          "(mean plausibility, % consistent).")


if __name__ == "__main__":
    main()
