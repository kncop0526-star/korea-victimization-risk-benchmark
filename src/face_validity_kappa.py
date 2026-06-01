"""
face_validity_kappa.py — inter-rater reliability for the §4.4 face-validity screen.

Reads two scored sheets (plausible_1to5 + consistent_Y_N per persona, same 50-row order) and reports:
per-rater mean/distribution, exact + within-1 agreement, quadratic-weighted Cohen's kappa on the
ordinal plausibility, and consistency-axis agreement. NumPy/openpyxl only.

Usage
  python src/face_validity_kappa.py rater1.xlsx rater2.xlsx
"""
from __future__ import annotations
import sys
import numpy as np
from openpyxl import load_workbook


def read_sheet(path):
    ws = load_workbook(path, data_only=True).active
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    pe = hdr.index("plausible_1to5") + 1
    cf = hdr.index("consistent_Y_N") + 1
    plaus, cons = [], []
    for r in range(2, ws.max_row + 1):
        p = ws.cell(r, pe).value
        c = ws.cell(r, cf).value
        if p is None and c is None:
            continue
        plaus.append(int(p)); cons.append(str(c).strip().upper())
    return np.array(plaus), cons


def qwk(a, b):
    levels = sorted(set(a.tolist() + b.tolist()))
    idx = {v: i for i, v in enumerate(levels)}; k = len(levels)
    O = np.zeros((k, k))
    for x, y in zip(a, b):
        O[idx[x], idx[y]] += 1
    O /= O.sum()
    row = O.sum(1, keepdims=True); col = O.sum(0, keepdims=True); E = row @ col
    W = np.array([[(levels[i] - levels[j]) ** 2 for j in range(k)] for i in range(k)], float)
    W /= (levels[-1] - levels[0]) ** 2
    den = float((W * E).sum())
    return float(1 - (W * O).sum() / den) if den > 0 else 1.0


def main():
    f1, f2 = sys.argv[1], sys.argv[2]
    a, ca = read_sheet(f1)
    b, cb = read_sheet(f2)
    n = min(len(a), len(b)); a, b = a[:n], b[:n]
    print(f"N = {n}")
    for lbl, v in [("rater A", a), ("rater B", b)]:
        dist = {lv: int((v == lv).sum()) for lv in range(1, 6)}
        print(f"  {lbl}: mean={v.mean():.3f}  >=4: {int((v>=4).sum())}/{n}  dist(1..5)={[dist[i] for i in range(1,6)]}")
    exact = float(np.mean(a == b)); within1 = float(np.mean(np.abs(a - b) <= 1))
    mae = float(np.mean(np.abs(a - b)))
    cons_agree = float(np.mean([x == y for x, y in zip(ca[:n], cb[:n])]))
    print(f"  plausibility: exact={exact:.3f}  within-1={within1:.3f}  MAE={mae:.3f}")
    print(f"  quadratic-weighted Cohen's kappa = {qwk(a, b):.4f}")
    print(f"  consistency-axis agreement = {cons_agree:.3f}  (raterA Y={ca.count('Y')}/{len(ca)}, raterB Y={cb.count('Y')}/{len(cb)})")


if __name__ == "__main__":
    main()
