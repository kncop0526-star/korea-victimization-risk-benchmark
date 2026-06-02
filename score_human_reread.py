#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_human_reread.py
---------------------------------------------------------------------------
Human re-read validation scorer for the persona-narrative dataset.

Given >=2 FILLED coder sheets and the original answer key, computes per
attribute:
  (1) Inter-coder agreement   : mean pairwise quadratic weighted kappa (QWK)
                                 + Krippendorff's alpha (ordinal)
  (2) Recovery of original     : QWK(consensus vs key) with 95% bootstrap CI
                                 + Spearman rho (robustness)
The consensus is the per-item mean of coders, rounded to the nearest integer
(half rounded up), clipped to 1..5.

Usage
-----
  py score_human_reread.py \
     --sheets human_reread_subset60_coder1_FILLED.xlsx \
              human_reread_subset60_coder2_FILLED.xlsx \
     --key results/human_reread_key.csv \
     --out results

Notes
-----
* Works with any number of coders (>=2).
* Key CSV is matched on persona_id; the 4 attribute columns are auto-detected
  by keyword (financial/vulnerab, digital/literac, authority/defer,
  social/isol). Override with --key-cols if auto-detection fails.
* Robust to blank cells (reported and skipped pairwise); fails loudly if a
  coder left an entire attribute empty.
"""
import argparse, sys, os, csv, math, re
from collections import OrderedDict

try:
    import numpy as np
except ImportError:
    sys.exit("numpy required: pip install numpy")
try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

ATTRS = OrderedDict([
    ("financial_vulnerability", ["financ", "vulnerab", "재정"]),
    ("digital_literacy",        ["digital", "literac", "디지털"]),
    ("authority_deference",     ["authorit", "defer", "권위"]),
    ("social_isolation",        ["social", "isolat", "고립"]),
])
SHEET_COLS = {  # column header in coder sheets
    "financial_vulnerability": "human_financial_vulnerability",
    "digital_literacy":        "human_digital_literacy",
    "authority_deference":     "human_authority_deference",
    "social_isolation":        "human_social_isolation",
}
ALPHA_THRESHOLD = 0.667   # README pre-registered acceptability line
SCALE_MIN, SCALE_MAX = 1, 5


# --------------------------------------------------------------------------- #
# metrics
# --------------------------------------------------------------------------- #
def quadratic_weighted_kappa(a, b, lo=SCALE_MIN, hi=SCALE_MAX):
    """QWK for two integer rating vectors on the ordinal scale lo..hi."""
    a = np.asarray(a, float); b = np.asarray(b, float)
    mask = ~(np.isnan(a) | np.isnan(b))
    a, b = a[mask], b[mask]
    if len(a) == 0:
        return float("nan")
    n_cat = hi - lo + 1
    O = np.zeros((n_cat, n_cat))
    for x, y in zip(a, b):
        O[int(round(x)) - lo, int(round(y)) - lo] += 1
    W = np.zeros((n_cat, n_cat))
    for i in range(n_cat):
        for j in range(n_cat):
            W[i, j] = (i - j) ** 2 / (n_cat - 1) ** 2
    act_a = O.sum(axis=1); act_b = O.sum(axis=0)
    E = np.outer(act_a, act_b) / O.sum()
    denom = (W * E).sum()
    if denom == 0:
        return float("nan")
    return 1.0 - (W * O).sum() / denom


def krippendorff_alpha_ordinal(matrix):
    """matrix: coders x items, NaN = missing. Ordinal alpha."""
    try:
        import krippendorff
        return float(krippendorff.alpha(reliability_data=matrix,
                                        level_of_measurement="ordinal"))
    except Exception:
        return _alpha_ordinal_fallback(matrix)


def _alpha_ordinal_fallback(matrix):
    """Self-contained ordinal Krippendorff alpha (no external dep)."""
    M = np.asarray(matrix, float)
    vals = M[~np.isnan(M)]
    if vals.size == 0:
        return float("nan")
    cats = sorted(set(int(round(v)) for v in vals))
    idx = {c: k for k, c in enumerate(cats)}
    nc = len(cats)
    # marginal counts of value pairs per unit
    n_g = np.zeros(nc)
    Do_num = 0.0; n_pairs_total = 0.0
    units = []
    for j in range(M.shape[1]):
        col = M[:, j]; col = col[~np.isnan(col)]
        if len(col) >= 2:
            units.append(col)
    # value frequencies (over pairable units)
    for col in units:
        for v in col:
            n_g[idx[int(round(v))]] += 1
    N = n_g.sum()
    if N < 2:
        return float("nan")

    def delta2(c1, c2):
        i, k = idx[c1], idx[c2]
        lo_, hi_ = (i, k) if i <= k else (k, i)
        s = sum(n_g[g] for g in range(lo_, hi_ + 1)) - (n_g[i] + n_g[k]) / 2.0
        return s ** 2

    # observed disagreement
    for col in units:
        m = len(col)
        for x in range(m):
            for y in range(m):
                if x == y:
                    continue
                Do_num += delta2(int(round(col[x])), int(round(col[y])))
            n_pairs_total += (m - 1)
    Do = Do_num / n_pairs_total if n_pairs_total else float("nan")
    # expected disagreement
    De_num = 0.0
    for c1 in cats:
        for c2 in cats:
            De_num += n_g[idx[c1]] * n_g[idx[c2]] * delta2(c1, c2)
    De = De_num / (N * (N - 1))
    if De == 0:
        return float("nan")
    return 1.0 - Do / De


def spearman_rho(a, b):
    a = np.asarray(a, float); b = np.asarray(b, float)
    mask = ~(np.isnan(a) | np.isnan(b)); a, b = a[mask], b[mask]
    if len(a) < 3:
        return float("nan")
    ra = _rankdata(a); rb = _rankdata(b)
    ra -= ra.mean(); rb -= rb.mean()
    d = math.sqrt((ra ** 2).sum() * (rb ** 2).sum())
    return float((ra * rb).sum() / d) if d else float("nan")


def _rankdata(x):
    order = np.argsort(x, kind="mergesort")
    ranks = np.empty(len(x), float); ranks[order] = np.arange(1, len(x) + 1)
    # average ties
    _, inv, cnt = np.unique(x, return_inverse=True, return_counts=True)
    csum = np.cumsum(cnt); start = csum - cnt
    avg = (start + csum - 1) / 2.0 + 1
    return avg[inv]


def consensus(matrix):
    """Per-item rounded mean across coders (half up), clipped to scale."""
    M = np.asarray(matrix, float)
    out = np.full(M.shape[1], np.nan)
    for j in range(M.shape[1]):
        col = M[:, j]; col = col[~np.isnan(col)]
        if len(col):
            out[j] = min(SCALE_MAX, max(SCALE_MIN,
                          int(math.floor(col.mean() + 0.5))))
    return out


# --------------------------------------------------------------------------- #
# IO
# --------------------------------------------------------------------------- #
def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["rater"] if "rater" in wb.sheetnames else wb[wb.sheetnames[-1]]
    header = [c.value for c in ws[1]]
    cidx = {h: i for i, h in enumerate(header)}
    if "persona_id" not in cidx:
        sys.exit(f"[{path}] missing 'persona_id' header")
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = row[cidx["persona_id"]]
        if pid in (None, "", "persona_id"):
            continue
        rec = {}
        for a, col in SHEET_COLS.items():
            v = row[cidx[col]] if col in cidx else None
            rec[a] = _to_score(v)
        data[pid] = rec
    return data


def _to_score(v):
    if v in (None, ""):
        return float("nan")
    try:
        f = float(v)
    except (TypeError, ValueError):
        return float("nan")
    return f


def read_key(path, overrides):
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        cols = reader.fieldnames or []
        pid_col = next((c for c in cols if re.search(r"persona.?id|^id$", c, re.I)), None)
        if pid_col is None:
            pid_col = cols[0]
        keymap = {}
        for a, kws in ATTRS.items():
            if a in overrides:
                keymap[a] = overrides[a]; continue
            hit = next((c for c in cols
                        if any(k in c.lower() for k in kws)), None)
            if hit is None:
                sys.exit(f"[key] cannot find column for '{a}'. "
                         f"Use --key-cols {a}=<colname>. Available: {cols}")
            keymap[a] = hit
        data = {}
        for row in reader:
            pid = row[pid_col]
            if not pid:
                continue
            data[pid] = {a: _to_score(row[keymap[a]]) for a in ATTRS}
    return data, keymap


# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--sheets", nargs="+", required=True,
                    help="2+ FILLED coder xlsx files")
    ap.add_argument("--key", required=True, help="original answer key CSV")
    ap.add_argument("--out", default=".", help="output directory")
    ap.add_argument("--key-cols", nargs="*", default=[],
                    help="manual key column map, e.g. financial_vulnerability=fv")
    ap.add_argument("--boot", type=int, default=2000, help="bootstrap iters")
    ap.add_argument("--seed", type=int, default=20240601)
    args = ap.parse_args()

    if len(args.sheets) < 2:
        sys.exit("need >=2 coder sheets for inter-coder reliability")
    overrides = dict(kv.split("=", 1) for kv in args.key_cols)

    coders = [read_sheet(p) for p in args.sheets]
    key, keymap = read_key(args.key, overrides)

    # common personas across all coders AND key
    ids = set(coders[0])
    for c in coders[1:]:
        ids &= set(c)
    ids &= set(key)
    ids = sorted(ids)
    if not ids:
        sys.exit("no overlapping persona_id across sheets and key")
    print(f"# coders={len(coders)}  overlapping personas={len(ids)}")
    print(f"# key columns: {keymap}\n")

    rng = np.random.default_rng(args.seed)
    rows_out = []
    print(f"{'attribute':24s} {'alpha':>7} {'meanQWK':>8} "
          f"{'recQWK':>7} {'CI95':>16} {'rho':>6}  flag")
    print("-" * 80)
    for a in ATTRS:
        # build coders x items matrix
        M = np.array([[c[i][a] for i in ids] for c in coders], float)
        # missing-cell guard
        if np.all(np.isnan(M[:, :]), axis=1).any():
            sys.exit(f"a coder left attribute '{a}' entirely blank")
        n_blank = int(np.isnan(M).sum())

        alpha = krippendorff_alpha_ordinal(M)
        # mean pairwise QWK among coders
        pair = []
        for x in range(len(coders)):
            for y in range(x + 1, len(coders)):
                pair.append(quadratic_weighted_kappa(M[x], M[y]))
        mean_qwk = float(np.nanmean(pair)) if pair else float("nan")

        cons = consensus(M)
        kv = np.array([key[i][a] for i in ids], float)
        rec = quadratic_weighted_kappa(cons, kv)
        rho = spearman_rho(cons, kv)

        # bootstrap CI on recovery QWK (resample personas)
        boots = []
        n = len(ids)
        for _ in range(args.boot):
            idx = rng.integers(0, n, n)
            boots.append(quadratic_weighted_kappa(cons[idx], kv[idx]))
        boots = np.array([b for b in boots if not math.isnan(b)])
        lo, hi = (np.percentile(boots, 2.5), np.percentile(boots, 97.5)) \
            if boots.size else (float("nan"), float("nan"))

        flag = "" if (not math.isnan(alpha) and alpha >= ALPHA_THRESHOLD) else "LOW-α"
        print(f"{a:24s} {alpha:7.3f} {mean_qwk:8.3f} {rec:7.3f} "
              f"[{lo:5.2f},{hi:5.2f}] {rho:6.3f}  {flag}")
        rows_out.append(dict(attribute=a, krippendorff_alpha=round(alpha, 4),
                             mean_pairwise_qwk=round(mean_qwk, 4),
                             recovery_qwk=round(rec, 4),
                             recovery_qwk_ci_low=round(lo, 4),
                             recovery_qwk_ci_high=round(hi, 4),
                             consensus_vs_key_spearman=round(rho, 4),
                             n_personas=len(ids), n_blank_cells=n_blank,
                             alpha_flag=flag or "ok"))

    os.makedirs(args.out, exist_ok=True)
    outp = os.path.join(args.out, "human_reread_results.csv")
    with open(outp, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows_out[0].keys()))
        w.writeheader(); w.writerows(rows_out)
    print("\nwrote", outp)
    print("\nInterpretation:")
    print("  recovery_qwk >= ~0.70  -> narratives carry the intended values")
    print(f"  krippendorff_alpha < {ALPHA_THRESHOLD} -> attribute ambiguous to humans")


if __name__ == "__main__":
    main()
